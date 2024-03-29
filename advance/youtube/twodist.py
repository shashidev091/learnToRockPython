import openpyxl as xl
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]

print(matrix)

for row in matrix:
    for item in row:
        if item == 5:
            item = 10
        print(item, end=', ')
    print('\n')

wb = xl.load_workbook('../../Financial Sample.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cellO = sheet.cell(1, 1)
print(cell.value)
print(cellO.value)
print(sheet.max_row)
print(sheet.max_column)
max_cols = sheet.max_column

for row in range(1, sheet.max_row + 1):
    cells = sheet.cell(row, 5)
    new_values = cells.value * 10
    # print(new_values)
    updated_cell = sheet.cell(row, max_cols)
    updated_cell.value = new_values

# wb.save('test.xlsx')


def process_workbook(filename):
    workbook = xl.load_workbook(filename)
    sheets = workbook['Sheet1']

    for rows in range(sheets.max_row + 1):
        sheet_cell = sheets.cell(rows, 5)
        updated_price = sheet_cell.value * 20 / 100
        updated_price_cell = sheets.cell(row, max_cols)
        updated_price_cell.value = updated_price

    workbook.save('anotherFile.xlsx')

"""
    - TODO
    Start flask with oops
"""